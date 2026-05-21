"""
《人文科學中心》課程表下載工具
Streamlit 網頁介面 - 供行政人員使用

啟動方式：
    streamlit run app.py
"""

import ssl
import re
import io
import urllib.request
import urllib.parse
import json

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
BASE_URL       = "https://timetable.nycu.edu.tw/?r=main/"
BOYA_UID       = "A91F7169-A56F-4612-85DE-2AC40893212C"
HUM_CENTER_UID = "F820DFDB-0D95-48BC-B9D9-F7939405AD83"


# ── API ───────────────────────────────────────────────────────────────────────
def api_post(endpoint: str, payload: dict):
    url  = BASE_URL + endpoint
    data = urllib.parse.urlencode(payload).encode("utf-8")
    req  = urllib.request.Request(url, data=data)
    ssl_context = ssl._create_unverified_context()
    with urllib.request.urlopen(req, timeout=30, context=ssl_context) as resp:
        return json.loads(resp.read().decode("utf-8"))


@st.cache_data(show_spinner=False)
def get_acysem_options():
    """回傳所有學年學期選項，格式：[('115 學年度 第1學期', '115', '1'), ...]"""
    acysem_list = api_post("get_acysem", {})
    options = []
    for item in acysem_list:
        t   = item["T"]           # e.g. "1151"
        acy = t[:-1]              # "115"
        sem = t[-1]               # "1"
        if sem == "X":
            label = f"{acy} 學年度 暑期"
        else:
            label = f"{acy} 學年度  第{'一' if sem=='1' else '二'}學期"
        options.append((label, acy, sem))
    return options


# ── 摘要解析 ──────────────────────────────────────────────────────────────────
def parse_brief(brief_text: str) -> dict:
    result = {}
    if not brief_text:
        return result
    for item in brief_text.split(","):
        item = item.strip()
        m = re.match(r'^(.+?)\((\d+)\)$', item)
        if m:
            cat  = m.group(1).strip()
            year = m.group(2).strip()
            result[year] = (result[year] + " / " + cat) if year in result else cat
    return result


# ── 抓課程 ───────────────────────────────────────────────────────────────────
def fetch_courses(acy: str, sem: str) -> list[dict]:
    raw = api_post("get_cos_list", {
        "m_acy": acy, "m_sem": sem, "m_acyend": acy, "m_semend": sem,
        "m_dep_uid": BOYA_UID,
        "m_group": "**", "m_grade": "**", "m_class": "**", "m_option": "**",
        "m_crsname": "**", "m_teaname": "**", "m_cos_id": "**",
        "m_cos_code": "**", "m_crstime": "**", "m_crsoutline": "**",
        "m_costype": "**", "m_selcampus": "**",
    })

    if HUM_CENTER_UID not in raw:
        raise RuntimeError("找不到「人文科學中心」資料，請稍後再試。")

    dept      = raw[HUM_CENTER_UID]
    brief_map = dept.get("brief", {})
    courses   = dept.get("1", {})

    rows = []
    for _, c in courses.items():
        cos_id_full = f"{acy}{sem}_{c.get('cos_id','')}"
        brief_info  = brief_map.get(cos_id_full, {})
        brief_text  = ""
        if isinstance(brief_info, dict):
            for sub in brief_info.values():
                if isinstance(sub, dict):
                    brief_text = sub.get("brief", "")
                    break

        parsed = parse_brief(brief_text)
        sem_label = "上" if sem == "1" else "下" if sem == "2" else sem

        rows.append({
            "學期別":                     f"{c.get('acy','')}{sem_label}",
            "課號":                       c.get("cos_id", ""),
            "永久課號":                   c.get("cos_code", ""),
            "114學年後使用類別\n(110制)": parsed.get("110", ""),
            "109學年(含)以前適用\n(106制)": parsed.get("106", ""),
            "舊制通識(90)":               parsed.get("90", ""),
            "舊制通識(96)":               parsed.get("96", ""),
            "摘要(原始)":                 brief_text,
            "課程名稱":                   c.get("cos_cname", ""),
            "英文課程名稱":               c.get("cos_ename", ""),
            "人數上限":                   c.get("num_limit", ""),
            "修課人數":                   c.get("reg_num", ""),
            "上課時間及教室":             c.get("cos_time", ""),
            "學分":                       c.get("cos_credit", ""),
            "時數":                       c.get("cos_hours", ""),
            "開課教師":                   c.get("teacher", ""),
            "選別":                       c.get("cos_type", ""),
            "備註":                       c.get("memo", ""),
        })
    return rows


# ── 產生 Excel（回傳 bytes） ──────────────────────────────────────────────────
def build_excel(rows: list[dict], acy: str, sem: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人文科學中心課程表"

    HEADER_FILL     = PatternFill("solid", fgColor="2E75B6")
    HEADER_FILL_110 = PatternFill("solid", fgColor="375623")
    HEADER_FILL_106 = PatternFill("solid", fgColor="7030A0")
    HEADER_FILL_OLD = PatternFill("solid", fgColor="808080")
    TITLE_FILL      = PatternFill("solid", fgColor="1F4E79")
    ALT_FILL        = PatternFill("solid", fgColor="D6E4F0")
    ALT_110_FILL    = PatternFill("solid", fgColor="E2EFDA")
    ALT_106_FILL    = PatternFill("solid", fgColor="EDE7F6")
    ALT_OLD_FILL    = PatternFill("solid", fgColor="F0F0F0")
    WHITE_FILL      = PatternFill("solid", fgColor="FFFFFF")

    HEADER_FONT = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=10)
    TITLE_FONT  = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=14)
    BODY_FONT   = Font(name="微軟正黑體", size=10)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="B0C4DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    columns   = list(rows[0].keys())
    col_count = len(columns)

    def hdr_fill(col):
        if "110" in col: return HEADER_FILL_110
        if "106" in col: return HEADER_FILL_106
        if "(90)" in col or "(96)" in col: return HEADER_FILL_OLD
        return HEADER_FILL

    def row_fill(col, alt):
        if "110" in col: return ALT_110_FILL if alt else WHITE_FILL
        if "106" in col: return ALT_106_FILL if alt else WHITE_FILL
        if "(90)" in col or "(96)" in col: return ALT_OLD_FILL if alt else WHITE_FILL
        return ALT_FILL if alt else WHITE_FILL

    sem_lbl    = "第一學期" if sem == "1" else "第二學期" if sem == "2" else f"第{sem}學期"
    title_text = f"國立陽明交通大學 {acy}學年度{sem_lbl} 《人文科學中心》課程表"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font = TITLE_FONT; tc.fill = TITLE_FILL; tc.alignment = CTR
    ws.row_dimensions[1].height = 32

    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = HEADER_FONT; c.fill = hdr_fill(col)
        c.alignment = CTR; c.border = border
    ws.row_dimensions[2].height = 44

    left_set = {"課程名稱","英文課程名稱","摘要(原始)","上課時間及教室",
                "開課教師","備註","114學年後使用類別\n(110制)",
                "109學年(含)以前適用\n(106制)","舊制通識(90)","舊制通識(96)"}
    for ri, row_data in enumerate(rows, 3):
        alt = (ri % 2 == 1)
        for ci, col in enumerate(columns, 1):
            c = ws.cell(row=ri, column=ci, value=row_data.get(col, ""))
            c.font = BODY_FONT; c.fill = row_fill(col, alt)
            c.alignment = LFT if col in left_set else CTR
            c.border = border
        ws.row_dimensions[ri].height = 18

    widths = {
        "學期別": 8, "課號": 9, "永久課號": 13,
        "114學年後使用類別\n(110制)": 26,
        "109學年(含)以前適用\n(106制)": 22,
        "舊制通識(90)": 20, "舊制通識(96)": 16,
        "摘要(原始)": 44, "課程名稱": 25, "英文課程名稱": 32,
        "人數上限": 9, "修課人數": 9, "上課時間及教室": 22,
        "學分": 7, "時數": 7, "開課教師": 12, "選別": 8, "備註": 22,
    }
    for ci, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 12)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(col_count)}2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════════════════════
#  Streamlit 介面
# ═════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="人文科學中心 課程表下載",
    page_icon="📚",
    layout="centered",
)

# ── 自訂 CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;500;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Noto Sans TC', sans-serif;
}

.main-card {
    background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 100%);
    border-radius: 16px;
    padding: 2rem 2.5rem 1.5rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 8px 32px rgba(31,78,121,0.25);
}
.main-card h1 {
    color: white;
    font-size: 1.7rem;
    font-weight: 700;
    margin: 0 0 0.3rem 0;
}
.main-card p {
    color: rgba(255,255,255,0.85);
    font-size: 0.95rem;
    margin: 0;
}

.info-box {
    background: #EFF6FF;
    border-left: 4px solid #2E75B6;
    border-radius: 8px;
    padding: 0.9rem 1.2rem;
    margin: 1rem 0;
    font-size: 0.9rem;
    color: #1e3a5f;
}

.legend {
    display: flex;
    gap: 1rem;
    flex-wrap: wrap;
    margin: 0.5rem 0 1rem;
}
.legend-item {
    display: flex;
    align-items: center;
    gap: 6px;
    font-size: 0.85rem;
    color: #444;
}
.dot {
    width: 14px; height: 14px;
    border-radius: 3px;
    display: inline-block;
}

/* 下載按鈕 */
.stDownloadButton > button {
    background: linear-gradient(135deg, #375623, #4e7a30) !important;
    color: white !important;
    font-size: 1.05rem !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.65rem 2rem !important;
    width: 100% !important;
    box-shadow: 0 4px 15px rgba(55,86,35,0.3) !important;
    transition: all 0.2s !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 20px rgba(55,86,35,0.4) !important;
}
</style>
""", unsafe_allow_html=True)

# ── 標題卡片 ──────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-card">
    <h1>📚 人文科學中心 課程表下載</h1>
    <p>國立陽明交通大學 · 自動從課程時間表系統抓取資料，匯出 Excel 檔案</p>
</div>
""", unsafe_allow_html=True)

# ── 說明 ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="info-box">
    ℹ️ &nbsp;下載的 Excel 會自動將「摘要」欄位拆分成以下獨立欄位，方便對照比較：<br><br>
    <div class="legend">
        <span class="legend-item"><span class="dot" style="background:#375623"></span>114學年後使用類別（110制）</span>
        <span class="legend-item"><span class="dot" style="background:#7030A0"></span>109學年(含)以前適用（106制）</span>
        <span class="legend-item"><span class="dot" style="background:#808080"></span>舊制通識（90 / 96）</span>
    </div>
</div>
""", unsafe_allow_html=True)

# ── 選學年度學期 ──────────────────────────────────────────────────────────────
with st.spinner("載入學年度清單…"):
    try:
        options = get_acysem_options()
    except Exception as e:
        st.error(f"❌ 無法連線至陽明交通大學課程系統：{e}")
        st.stop()

labels   = [o[0] for o in options]
selected = st.selectbox("📅  請選擇學年度學期", labels, index=0)
chosen   = next(o for o in options if o[0] == selected)
acy, sem = chosen[1], chosen[2]

st.markdown(f"已選擇：**{selected.strip()}**")

# ── 產生並下載 ────────────────────────────────────────────────────────────────
if st.button("🔍  預覽 & 準備下載", use_container_width=True):
    with st.spinner("正在從課程時間表抓取資料，請稍候…"):
        try:
            rows = fetch_courses(acy, sem)
            excel_bytes = build_excel(rows, acy, sem)
            st.session_state["rows"]        = rows
            st.session_state["excel_bytes"] = excel_bytes
            st.session_state["acy"]         = acy
            st.session_state["sem"]         = sem
            st.success(f"✅ 已找到 **{len(rows)}** 筆課程！")
        except Exception as e:
            st.error(f"❌ 發生錯誤：{e}")

# ── 顯示預覽表格 + 下載按鈕 ──────────────────────────────────────────────────
if "rows" in st.session_state:
    rows        = st.session_state["rows"]
    excel_bytes = st.session_state["excel_bytes"]
    acy_s       = st.session_state["acy"]
    sem_s       = st.session_state["sem"]
    sem_lbl     = "第一學期" if sem_s == "1" else "第二學期" if sem_s == "2" else f"第{sem_s}學期"
    filename    = f"人文科學中心課程表_{acy_s}學年度{sem_lbl}.xlsx"

    # 下載按鈕
    st.download_button(
        label=f"⬇️  下載 Excel（{len(rows)} 筆）",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # 預覽表格
    st.markdown("#### 📋 資料預覽（前 10 筆）")
    import pandas as pd
    df = pd.DataFrame(rows)
    preview_cols = ["學期別","永久課號","課程名稱",
                    "114學年後使用類別\n(110制)",
                    "109學年(含)以前適用\n(106制)",
                    "開課教師","上課時間及教室","選別"]
    st.dataframe(
        df[preview_cols].head(10).rename(columns={
            "114學年後使用類別\n(110制)": "110制",
            "109學年(含)以前適用\n(106制)": "106制",
        }),
        use_container_width=True,
        hide_index=True,
    )

    st.caption(f"資料來源：國立陽明交通大學課程時間表 https://timetable.nycu.edu.tw/")

# ── 頁尾 ──────────────────────────────────────────────────────────────────────
st.markdown("---")
st.caption("本工具自動從陽明交通大學課程時間表系統擷取《人文科學中心》課程資料。如遇問題請洽系統管理員。")
