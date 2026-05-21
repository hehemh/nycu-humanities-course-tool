"""
國立陽明交通大學《人文科學中心》課程表爬蟲
抓取指定學年度學期的課程資料，匯出成 Excel 檔案

用法：
    python nycu_humanities_courses.py
    python nycu_humanities_courses.py --acy 115 --sem 1
"""

import re
import urllib.request
import urllib.parse
import json
import argparse

# ── 嘗試匯入 openpyxl，不存在則提示安裝 ──────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少 openpyxl，請執行：pip install openpyxl")
    raise

BASE_URL = "https://timetable.nycu.edu.tw/?r=main/"

# UIDs（固定，不會改變）
TYPE_OTHER_UID = "94EAAC09-C3BC-4BEB-BB9B-47E6F5F652C8"   # 其他課程
BOYA_UID       = "A91F7169-A56F-4612-85DE-2AC40893212C"   # 博雅書苑
HUM_CENTER_UID = "F820DFDB-0D95-48BC-B9D9-F7939405AD83"   # 人文科學中心


def api_post(endpoint: str, payload: dict) -> dict | list:
    """向 NYCU 課程 API 發送 POST 請求並回傳 JSON。"""
    url  = BASE_URL + endpoint
    data = urllib.parse.urlencode(payload).encode("utf-8")
    req  = urllib.request.Request(url, data=data)
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def get_latest_acysem() -> tuple[str, str]:
    """取得最新的學年度與學期。"""
    acysem_list = api_post("get_acysem", {})
    latest = acysem_list[0]["T"]   # e.g. "1151"
    acy = latest[:-1]              # "115"
    sem = latest[-1]               # "1"
    return acy, sem


def parse_brief(brief_text: str) -> dict[str, str]:
    """
    將摘要字串拆解成各年份制度的分類。

    輸入範例：
        "文化/基礎(96),核心-人文(106),核心通識-藝術與文化(90),領域課程-人文與美學(110)"

    回傳：
        {
            "110": "領域課程-人文與美學",
            "106": "核心-人文",
            "90":  "核心通識-藝術與文化",
            "96":  "文化/基礎",
        }
    """
    result = {}
    if not brief_text:
        return result

    # 每個項目格式為 "名稱(年份)"，用逗號分隔
    for item in brief_text.split(","):
        item = item.strip()
        m = re.match(r'^(.+?)\((\d+)\)$', item)
        if m:
            category = m.group(1).strip()
            year     = m.group(2).strip()
            # 若同一年份有多個分類（罕見），用「/」連接
            if year in result:
                result[year] = result[year] + " / " + category
            else:
                result[year] = category

    return result


def fetch_humanities_courses(acy: str, sem: str) -> list[dict]:
    """
    抓取《人文科學中心》所有課程資料。
    回傳欄位清單，每筆為一門課。
    """
    print(f"正在抓取 {acy} 學年度第 {sem} 學期《人文科學中心》課程表…")

    raw = api_post("get_cos_list", {
        "m_acy":        acy,
        "m_sem":        sem,
        "m_acyend":     acy,
        "m_semend":     sem,
        "m_dep_uid":    BOYA_UID,
        "m_group":      "**",
        "m_grade":      "**",
        "m_class":      "**",
        "m_option":     "**",
        "m_crsname":    "**",
        "m_teaname":    "**",
        "m_cos_id":     "**",
        "m_cos_code":   "**",
        "m_crstime":    "**",
        "m_crsoutline": "**",
        "m_costype":    "**",
        "m_selcampus":  "**",
    })

    if HUM_CENTER_UID not in raw:
        raise RuntimeError("在回應中找不到「人文科學中心」資料，請確認 UID 是否正確。")

    dept_data = raw[HUM_CENTER_UID]

    # brief_map: { "1151_561000": { "brief_code_str": { "brief": "...", ... } } }
    brief_map: dict = dept_data.get("brief", {})

    # 課程主資料在 key="1" 下
    courses_raw: dict = dept_data.get("1", {})

    rows = []
    for key, c in courses_raw.items():
        # ── 取得 brief 完整文字 ────────────────────────────────────────────────
        cos_id_full = f"{acy}{sem}_{c.get('cos_id', '')}"
        brief_info  = brief_map.get(cos_id_full, {})
        brief_text  = ""
        if isinstance(brief_info, dict):
            for sub in brief_info.values():
                if isinstance(sub, dict):
                    brief_text = sub.get("brief", "")
                    break

        # ── 拆解 brief 成各年份欄位 ──────────────────────────────────────────
        parsed = parse_brief(brief_text)

        row = {
            # ── 基本欄位 ────────────────────────────────────────────────────
            "學期別":                          f"{c.get('acy','')}{'上' if c.get('sem','')=='1' else '下' if c.get('sem','')=='2' else c.get('sem','')}",
            "課號":                            c.get("cos_id", ""),
            "永久課號":                        c.get("cos_code", ""),
            # ── 摘要拆分欄位（主要） ────────────────────────────────────────
            "114學年後使用類別(110)":           parsed.get("110", ""),
            "(109學年含以前適用)(106)":         parsed.get("106", ""),
            # ── 舊制度（較少使用，保留供參考） ────────────────────────────
            "舊制通識(90)":                    parsed.get("90", ""),
            "舊制通識(96)":                    parsed.get("96", ""),
            # ── 摘要原始文字（完整備查） ────────────────────────────────────
            "摘要(原始)":                      brief_text,
            # ── 其他課程資訊 ────────────────────────────────────────────────
            "課程名稱":                        c.get("cos_cname", ""),
            "英文課程名稱":                    c.get("cos_ename", ""),
            "人數上限":                        c.get("num_limit", ""),
            "修課人數":                        c.get("reg_num", ""),
            "上課時間及教室":                  c.get("cos_time", ""),
            "學分":                            c.get("cos_credit", ""),
            "時數":                            c.get("cos_hours", ""),
            "開課教師":                        c.get("teacher", ""),
            "選別":                            c.get("cos_type", ""),
            "備註":                            c.get("memo", ""),
        }
        rows.append(row)

    print(f"共找到 {len(rows)} 筆課程。")
    return rows


def export_to_excel(rows: list[dict], acy: str, sem: str, output_path: str):
    """將課程資料寫入格式化的 Excel 檔案。"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人文科學中心課程表"

    # ── 顏色定義 ──────────────────────────────────────────────────────────────
    HEADER_FILL      = PatternFill("solid", fgColor="2E75B6")   # 深藍（一般欄）
    HEADER_FILL_110  = PatternFill("solid", fgColor="375623")   # 深綠（110 制度）
    HEADER_FILL_106  = PatternFill("solid", fgColor="7030A0")   # 紫（106 制度）
    HEADER_FILL_OLD  = PatternFill("solid", fgColor="808080")   # 灰（舊制）
    TITLE_FILL       = PatternFill("solid", fgColor="1F4E79")   # 更深藍
    ALT_ROW_FILL     = PatternFill("solid", fgColor="D6E4F0")   # 淡藍（交替列）
    WHITE_FILL       = PatternFill("solid", fgColor="FFFFFF")
    # 摘要欄交替列顏色
    ALT_110_FILL     = PatternFill("solid", fgColor="E2EFDA")   # 淡綠
    ALT_106_FILL     = PatternFill("solid", fgColor="EDE7F6")   # 淡紫
    ALT_OLD_FILL     = PatternFill("solid", fgColor="F0F0F0")   # 淡灰

    HEADER_FONT      = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=10)
    TITLE_FONT       = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=14)
    BODY_FONT        = Font(name="微軟正黑體", size=10)
    CENTER_ALIGN     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGN       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin_side   = Side(style="thin",   color="B0C4DE")
    medium_side = Side(style="medium", color="2E75B6")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    columns = list(rows[0].keys()) if rows else []
    col_count = len(columns)

    # ── 欄標題對應的 header fill（依欄名決定顏色）────────────────────────────
    def header_fill_for(col_name: str) -> PatternFill:
        if "(110)" in col_name:
            return HEADER_FILL_110
        if "(106)" in col_name:
            return HEADER_FILL_106
        if "(90)" in col_name or "(96)" in col_name:
            return HEADER_FILL_OLD
        return HEADER_FILL

    # ── 資料列對應的 cell fill（依欄名 + 奇偶列決定顏色）─────────────────────
    def cell_fill_for(col_name: str, is_alt_row: bool) -> PatternFill:
        if "(110)" in col_name:
            return ALT_110_FILL if is_alt_row else WHITE_FILL
        if "(106)" in col_name:
            return ALT_106_FILL if is_alt_row else WHITE_FILL
        if "(90)" in col_name or "(96)" in col_name:
            return ALT_OLD_FILL if is_alt_row else WHITE_FILL
        return ALT_ROW_FILL if is_alt_row else WHITE_FILL

    # ── 標題列（第1行） ───────────────────────────────────────────────────────
    sem_label  = "第一學期" if sem == "1" else "第二學期" if sem == "2" else f"第{sem}學期"
    title_text = f"國立陽明交通大學 {acy}學年度{sem_label} 《人文科學中心》課程表"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell           = ws.cell(row=1, column=1, value=title_text)
    title_cell.font      = TITLE_FONT
    title_cell.fill      = TITLE_FILL
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 32

    # ── 欄標題（第2行） ───────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, start=1):
        cell           = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = header_fill_for(col_name)
        cell.alignment = CENTER_ALIGN
        cell.border    = thin_border
    ws.row_dimensions[2].height = 40   # 換行顯示較高

    # ── 資料列 ────────────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=3):
        is_alt = (row_idx % 2 == 1)
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = BODY_FONT
            cell.fill   = cell_fill_for(col_name, is_alt)
            cell.border = thin_border
            # 對齊方式
            left_cols = {"課程名稱", "英文課程名稱", "摘要(原始)",
                         "上課時間及教室", "開課教師", "備註",
                         "114學年後使用類別(110)", "(109學年含以前適用)(106)",
                         "舊制通識(90)", "舊制通識(96)"}
            cell.alignment = LEFT_ALIGN if col_name in left_cols else CENTER_ALIGN
        ws.row_dimensions[row_idx].height = 18

    # ── 欄寬設定 ──────────────────────────────────────────────────────────────
    col_widths = {
        "學期別":                   8,
        "課號":                     9,
        "永久課號":                 13,
        "114學年後使用類別(110)":   28,
        "(109學年含以前適用)(106)": 22,
        "舊制通識(90)":             22,
        "舊制通識(96)":             18,
        "摘要(原始)":               45,
        "課程名稱":                 25,
        "英文課程名稱":             32,
        "人數上限":                  9,
        "修課人數":                  9,
        "上課時間及教室":           22,
        "學分":                      7,
        "時數":                      7,
        "開課教師":                 12,
        "選別":                      8,
        "備註":                     22,
    }
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 12)

    # ── 凍結窗格 ──────────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── 自動篩選 ──────────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{get_column_letter(col_count)}2"

    wb.save(output_path)
    print(f"\n✅ 已匯出到：{output_path}")


def main():
    parser = argparse.ArgumentParser(description="抓取 NYCU 人文科學中心課程表並匯出 Excel")
    parser.add_argument("--acy",    type=str, default=None, help="學年度（例如 115），預設抓最新學期")
    parser.add_argument("--sem",    type=str, default=None, help="學期（1 或 2），預設抓最新學期")
    parser.add_argument("--output", type=str, default=None, help="輸出檔名（.xlsx）")
    args = parser.parse_args()

    # 若未指定學年度/學期，自動取最新
    if args.acy and args.sem:
        acy, sem = args.acy, args.sem
    else:
        print("未指定學年度/學期，自動取得最新資料…")
        acy, sem = get_latest_acysem()
        print(f"最新學期：{acy} 學年度第 {sem} 學期")

    output_path = args.output or f"人文科學中心課程表_{acy}學年度第{sem}學期.xlsx"

    rows = fetch_humanities_courses(acy, sem)

    if not rows:
        print("⚠️  沒有找到任何課程資料。")
        return

    export_to_excel(rows, acy, sem, output_path)

    # 預覽摘要拆分效果
    print("\n📋 摘要拆分預覽（前 5 筆）：")
    print(f"  {'永久課號':<13} {'課程名稱':<20} {'110制':<28} {'106制':<22}")
    print("  " + "-" * 85)
    for r in rows[:5]:
        print(f"  {r['永久課號']:<13} {r['課程名稱']:<20} {r['114學年後使用類別(110)']:<28} {r['(109學年含以前適用)(106)']:<22}")
    print(f"\n  共 {len(rows)} 筆課程")


if __name__ == "__main__":
    main()
